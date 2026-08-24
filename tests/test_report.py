from pathlib import Path

from openpyxl import load_workbook

from build_report import build_workbook, calculate_metrics, load_rows, summarize_metrics


def test_full_year_dataset_is_loaded():
    rows = load_rows()
    assert len(rows) == 12
    assert rows[0]["month"] == "2026-01"
    assert rows[-1]["month"] == "2026-12"


def test_metrics_are_calculated():
    metrics = calculate_metrics(load_rows())
    assert metrics[0]["profit"] == 7400
    assert round(metrics[0]["revenue_variance_pct"], 4) == 0.04
    assert metrics[0]["revenue_growth_pct"] == 0
    assert round(metrics[2]["rolling_3m_profit"], 2) == 7550.00


def test_summary_metrics_match_expected_totals():
    summary = summarize_metrics(calculate_metrics(load_rows()))
    assert summary["total_revenue"] == 267600
    assert summary["total_cost"] == 151600
    assert summary["total_profit"] == 116000
    assert summary["revenue_variance"] == 11100
    assert round(summary["profit_margin"], 4) == 0.4335


def test_workbook_is_generated_with_expected_sheets(tmp_path: Path):
    output = tmp_path / "report.xlsx"
    path, metrics = build_workbook(load_rows(), output)

    assert path.exists()
    assert path.stat().st_size > 0
    assert len(metrics) == 12

    workbook = load_workbook(path)
    assert workbook.sheetnames == ["Raw Data", "Analysis", "Dashboard"]
    assert workbook.active.title == "Dashboard"


def test_dashboard_contains_kpis_and_charts(tmp_path: Path):
    output = tmp_path / "report.xlsx"
    path, _ = build_workbook(load_rows(), output)
    workbook = load_workbook(path)
    dashboard = workbook["Dashboard"]

    assert dashboard["A5"].value == 267600
    assert dashboard["G5"].value == 116000
    assert round(dashboard["J5"].value, 4) == 0.4335
    assert len(dashboard._charts) == 4
