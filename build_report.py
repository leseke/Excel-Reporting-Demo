from __future__ import annotations

from csv import DictReader
from pathlib import Path
from statistics import mean

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parent
SOURCE = ROOT / "data" / "monthly_finance.csv"
OUTPUT = ROOT / "output" / "business_report.xlsx"

NAVY = "17324D"
TERRACOTTA = "B75B3E"
CREAM = "F7F1E8"
PALE_GREEN = "E8F3EC"
PALE_RED = "FCE8E5"
WHITE = "FFFFFF"
DARK = "1F2933"
BORDER = "D8D0C4"
MUTED = "667085"


def load_rows(path: Path = SOURCE):
    with path.open(encoding="utf-8", newline="") as f:
        return list(DictReader(f))


def calculate_metrics(rows):
    result, profits, previous_revenue = [], [], None
    for row in rows:
        revenue = float(row["revenue"]); budget_revenue = float(row["budget_revenue"])
        cost = float(row["cost"]); budget_cost = float(row["budget_cost"])
        profit = revenue - cost; profits.append(profit)
        result.append({
            "month": row["month"], "revenue": revenue, "budget_revenue": budget_revenue,
            "revenue_variance": revenue-budget_revenue,
            "revenue_variance_pct": (revenue-budget_revenue)/budget_revenue if budget_revenue else 0,
            "cost": cost, "budget_cost": budget_cost, "cost_variance": cost-budget_cost,
            "profit": profit, "profit_margin": profit/revenue if revenue else 0,
            "revenue_growth_pct": ((revenue-previous_revenue)/previous_revenue) if previous_revenue else 0,
            "rolling_3m_profit": mean(profits[-3:]),
        })
        previous_revenue = revenue
    return result


def summarize_metrics(metrics):
    tr=sum(x["revenue"] for x in metrics); tbr=sum(x["budget_revenue"] for x in metrics)
    tc=sum(x["cost"] for x in metrics); tbc=sum(x["budget_cost"] for x in metrics); tp=tr-tc
    return {"total_revenue":tr,"total_budget_revenue":tbr,"revenue_variance":tr-tbr,
            "revenue_variance_pct":(tr-tbr)/tbr if tbr else 0,"total_cost":tc,"total_budget_cost":tbc,
            "cost_variance":tc-tbc,"total_profit":tp,"profit_margin":tp/tr if tr else 0,
            "latest_revenue_growth_pct":metrics[-1]["revenue_growth_pct"] if metrics else 0}


def style_header(row):
    for cell in row:
        cell.fill=PatternFill("solid",fgColor=NAVY); cell.font=Font(color=WHITE,bold=True)
        cell.alignment=Alignment(horizontal="center")


def fit_columns(sheet, minimum=11, maximum=24):
    for cells in sheet.columns:
        length=max(len(str(c.value)) if c.value is not None else 0 for c in cells)
        sheet.column_dimensions[get_column_letter(cells[0].column)].width=min(max(length+2,minimum),maximum)


def add_table(sheet, ref, name):
    table=Table(displayName=name,ref=ref)
    table.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showFirstColumn=False,showLastColumn=False,showRowStripes=True,showColumnStripes=False)
    sheet.add_table(table)


def add_kpi_card(sheet,label_cell,value_cell,label,value,number_format):
    label_ref=sheet[label_cell]; value_ref=sheet[value_cell]
    label_ref.value=label; value_ref.value=value
    label_ref.fill=PatternFill("solid",fgColor=NAVY); label_ref.font=Font(color=WHITE,bold=True,size=11)
    value_ref.fill=PatternFill("solid",fgColor=CREAM); value_ref.font=Font(color=DARK,bold=True,size=16)
    value_ref.number_format=number_format
    for cell in (label_ref,value_ref):
        cell.alignment=Alignment(horizontal="center",vertical="center")
        cell.border=Border(*( [Side(style="thin",color=BORDER)]*0 )) if False else Border(left=Side(style="thin",color=BORDER),right=Side(style="thin",color=BORDER),top=Side(style="thin",color=BORDER),bottom=Side(style="thin",color=BORDER))


def build_workbook(rows, output_path: Path = OUTPUT):
    metrics=calculate_metrics(rows); summary=summarize_metrics(metrics); output_path.parent.mkdir(parents=True,exist_ok=True)
    wb=Workbook()

    start=wb.active; start.title="START HERE"; start.sheet_view.showGridLines=False
    start.merge_cells("A1:H1"); start["A1"]="EXCEL REPORTING DEMO"; start["A1"].fill=PatternFill("solid",fgColor=NAVY); start["A1"].font=Font(color=WHITE,bold=True,size=22); start["A1"].alignment=Alignment(horizontal="center")
    start.merge_cells("A2:H2"); start["A2"]="Client-ready automated reporting workbook"; start["A2"].font=Font(italic=True,color=MUTED); start["A2"].alignment=Alignment(horizontal="center")
    start.append([]); start.append(["QUICK START",""])
    instructions=["Replace the source CSV with the same column structure.","Run python build_report.py to rebuild the workbook.","Review Data Quality before using the report.","Open Dashboard for management KPIs and trends.","Use Analysis for detailed monthly calculations."]
    for i,text in enumerate(instructions,1): start.append([i,text])
    start["D4"]="WORKBOOK MAP"; start["D5"]="Raw Data"; start["E5"]="Auditable source layer"; start["D6"]="Analysis"; start["E6"]="Calculated reporting layer"; start["D7"]="Dashboard"; start["E7"]="Decision-ready KPIs and charts"; start["D8"]="Data Quality"; start["E8"]="Pre-delivery validation"
    for cell in (start["A4"],start["D4"]): cell.fill=PatternFill("solid",fgColor=TERRACOTTA); cell.font=Font(color=WHITE,bold=True)
    start.column_dimensions["A"].width=12; start.column_dimensions["B"].width=52; start.column_dimensions["D"].width=18; start.column_dimensions["E"].width=34

    raw=wb.create_sheet("Raw Data"); raw.sheet_view.showGridLines=False; raw.freeze_panes="A2"
    raw.append(["Month","Revenue","Budget Revenue","Cost","Budget Cost"])
    for row in rows: raw.append([row["month"],float(row["revenue"]),float(row["budget_revenue"]),float(row["cost"]),float(row["budget_cost"])])
    style_header(raw[1]); add_table(raw,f"A1:E{raw.max_row}","RawFinanceData"); fit_columns(raw)
    for r in range(2,raw.max_row+1):
        for c in range(2,6): raw.cell(r,c).number_format='#,##0.00 [$€-fr-FR]'

    analysis=wb.create_sheet("Analysis"); analysis.sheet_view.showGridLines=False; analysis.freeze_panes="A2"
    analysis.append(["Month","Revenue","Budget Revenue","Revenue Variance","Revenue Variance %","Cost","Budget Cost","Cost Variance","Profit","Profit Margin","Revenue Growth %","Rolling 3M Profit"])
    for x in metrics: analysis.append([x[k] for k in ["month","revenue","budget_revenue","revenue_variance","revenue_variance_pct","cost","budget_cost","cost_variance","profit","profit_margin","revenue_growth_pct","rolling_3m_profit"]])
    style_header(analysis[1]); add_table(analysis,f"A1:L{analysis.max_row}","ReportingAnalysis"); fit_columns(analysis,maximum=20)
    for r in range(2,analysis.max_row+1):
        for c in (2,3,4,6,7,8,9,12): analysis.cell(r,c).number_format='#,##0.00 [$€-fr-FR]'
        for c in (5,10,11): analysis.cell(r,c).number_format="0.0%"
    analysis.conditional_formatting.add(f"D2:D{analysis.max_row}",CellIsRule(operator="greaterThanOrEqual",formula=["0"],fill=PatternFill("solid",fgColor=PALE_GREEN)))
    analysis.conditional_formatting.add(f"D2:D{analysis.max_row}",CellIsRule(operator="lessThan",formula=["0"],fill=PatternFill("solid",fgColor=PALE_RED)))

    dashboard=wb.create_sheet("Dashboard"); dashboard.sheet_view.showGridLines=False
    dashboard["A1"]="Business Performance Dashboard"; dashboard["A1"].font=Font(bold=True,size=22,color=NAVY); dashboard.merge_cells("A1:P1")
    dashboard["A2"]="Synthetic 2026 management reporting demo — generated automatically from CSV source data"; dashboard["A2"].font=Font(italic=True,color=MUTED); dashboard.merge_cells("A2:P2")
    cards=[("A4","A5","TOTAL REVENUE",summary["total_revenue"],'#,##0 [$€-fr-FR]'),("D4","D5","TOTAL COST",summary["total_cost"],'#,##0 [$€-fr-FR]'),("G4","G5","NET PROFIT",summary["total_profit"],'#,##0 [$€-fr-FR]'),("J4","J5","PROFIT MARGIN",summary["profit_margin"],"0.0%"),("M4","M5","REVENUE VS BUDGET",summary["revenue_variance_pct"],"0.0%")]
    for args in cards: add_kpi_card(dashboard,*args)
    dashboard["A7"]="Management view"; dashboard["A7"].font=Font(bold=True,size=13,color=TERRACOTTA)
    dashboard["A8"]="Revenue variance"; dashboard["B8"]=summary["revenue_variance"]; dashboard["B8"].number_format='#,##0 [$€-fr-FR]'
    dashboard["D8"]="Cost variance"; dashboard["E8"]=summary["cost_variance"]; dashboard["E8"].number_format='#,##0 [$€-fr-FR]'
    dashboard["G8"]="Latest revenue growth"; dashboard["H8"]=summary["latest_revenue_growth_pct"]; dashboard["H8"].number_format="0.0%"
    line=LineChart(); line.title="Revenue vs Budget"; line.style=13; line.height=8; line.width=15; line.add_data(Reference(analysis,min_col=2,max_col=3,min_row=1,max_row=analysis.max_row),titles_from_data=True); line.set_categories(Reference(analysis,min_col=1,min_row=2,max_row=analysis.max_row)); dashboard.add_chart(line,"A10")
    bar=BarChart(); bar.title="Monthly Net Profit"; bar.style=10; bar.height=8; bar.width=15; bar.add_data(Reference(analysis,min_col=9,min_row=1,max_row=analysis.max_row),titles_from_data=True); bar.set_categories(Reference(analysis,min_col=1,min_row=2,max_row=analysis.max_row)); dashboard.add_chart(bar,"I10")
    growth=LineChart(); growth.title="Revenue Growth %"; growth.style=12; growth.height=7; growth.width=15; growth.add_data(Reference(analysis,min_col=11,min_row=1,max_row=analysis.max_row),titles_from_data=True); growth.set_categories(Reference(analysis,min_col=1,min_row=2,max_row=analysis.max_row)); dashboard.add_chart(growth,"A26")
    rolling=LineChart(); rolling.title="Rolling 3-Month Profit"; rolling.style=13; rolling.height=7; rolling.width=15; rolling.add_data(Reference(analysis,min_col=12,min_row=1,max_row=analysis.max_row),titles_from_data=True); rolling.set_categories(Reference(analysis,min_col=1,min_row=2,max_row=analysis.max_row)); dashboard.add_chart(rolling,"I26")
    for c in range(1,17): dashboard.column_dimensions[get_column_letter(c)].width=12

    quality=wb.create_sheet("Data Quality"); quality.sheet_view.showGridLines=False
    quality.merge_cells("A1:F1"); quality["A1"]="DATA QUALITY CONTROL"; quality["A1"].fill=PatternFill("solid",fgColor=NAVY); quality["A1"].font=Font(color=WHITE,bold=True,size=20); quality["A1"].alignment=Alignment(horizontal="center")
    quality.append([]); quality.append([]); quality.append(["Check","Result","Status"]); style_header(quality[4])
    checks=[("Rows loaded",len(rows),len(rows)>0),("Blank months",sum(not r.get("month") for r in rows),all(r.get("month") for r in rows)),("Zero revenue",sum(float(r["revenue"])<=0 for r in rows),all(float(r["revenue"])>0 for r in rows)),("Zero budget revenue",sum(float(r["budget_revenue"])<=0 for r in rows),all(float(r["budget_revenue"])>0 for r in rows)),("Margin outside 0–100%",sum(not 0<=x["profit_margin"]<=1 for x in metrics),all(0<=x["profit_margin"]<=1 for x in metrics))]
    for name,value,ok in checks: quality.append([name,value,"OK" if ok else "CHECK"])
    quality.column_dimensions["A"].width=28; quality.column_dimensions["B"].width=14; quality.column_dimensions["C"].width=14
    quality.conditional_formatting.add(f"C5:C{quality.max_row}",CellIsRule(operator="equal",formula=['"OK"'],fill=PatternFill("solid",fgColor=PALE_GREEN)))
    quality.conditional_formatting.add(f"C5:C{quality.max_row}",CellIsRule(operator="equal",formula=['"CHECK"'],fill=PatternFill("solid",fgColor=PALE_RED)))

    wb.active=wb.sheetnames.index("START HERE"); wb.save(output_path); return output_path,metrics


if __name__=="__main__":
    path,metrics=build_workbook(load_rows()); summary=summarize_metrics(metrics)
    print(f"Created {path} with {len(metrics)} reporting periods")
    print(f"Revenue: €{summary['total_revenue']:,.0f} | Profit: €{summary['total_profit']:,.0f} | Margin: {summary['profit_margin']:.1%}")
